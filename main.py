# -*- coding: utf-8 -*-
"""
小学生にもわかる説明：
  「layout.json」という設計図を読みこんで、画面の部品(ラベルや入力など)を
  決められた場所に並べます。Excel(xlsm)から読み書きもできます。
"""

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtGui import QIntValidator, QDoubleValidator, QRegularExpressionValidator
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPlainTextEdit, QPushButton, QFileDialog,
    QScrollArea, QFrame, QStatusBar, QMessageBox, QComboBox,
    QHBoxLayout, QListView, QStyledItemDelegate
)
from qt_material import apply_stylesheet
from openpyxl import load_workbook

import sys
import json
import os
import re
import logging  # ログ出力を扱う標準ライブラリです
from typing import Dict, Optional, Any, List, Callable
import threading
import time  # 時間を測るためのモジュールです
import importlib  # 追加のモジュールを読み込むための道具です

# 半角数字を全角数字に直すためのテーブルを用意します
_FW_TABLE = str.maketrans("0123456789", "０１２３４５６７８９")

# ログファイルを置く場所と名前を決めます
_LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
_LOG_FILE = os.path.join(_LOG_DIR, "application.log")


def _setup_logger() -> logging.Logger:
    """
    小学生にもわかる説明：
      ログを書き込む道具を準備して、画面とファイルの両方へ記録します。
    """
    # 初学者向け説明：最初に保存フォルダーが無ければ作ります。
    os.makedirs(_LOG_DIR, exist_ok=True)

    # 初学者向け説明：同じ設定を繰り返さないよう、すでに作成済みか確認します。
    logger = logging.getLogger("cdata")
    if logger.handlers:
        return logger

    # 初学者向け説明：INFO 以上の重要度を持つログを扱うように設定します。
    logger.setLevel(logging.INFO)

    # 初学者向け説明：ログの表示方法を統一するフォーマッターを用意します。
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s"
    )

    # 初学者向け説明：ログをファイルに書き出すハンドラーです。
    file_handler = logging.FileHandler(_LOG_FILE, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # 初学者向け説明：同じ内容を画面にも表示するハンドラーを追加します。
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger


# 初学者向け説明：アプリ全体で使う共通のロガーを作成します。
LOGGER = _setup_logger()

# 検索に利用する列名です。必要に応じて増やしてください。
SEARCH_COLUMNS = ["品目番号"]


def _summarize_for_log(data: Dict[str, str], max_length: int = 60) -> Dict[str, str]:
    """
    小学生にもわかる説明：
      長い文字は見やすい長さに切りそろえて、ログに残しやすくします。
    """
    # 初学者向け説明：新しい辞書を用意して、各値の長さを調整して入れ直します。
    summarized: Dict[str, str] = {}
    for key, value in data.items():
        text = "" if value is None else str(value)
        if len(text) > max_length:
            text = text[:max_length - 3] + "..."
        summarized[key] = text
    return summarized


def normalize_header_name(value: Any) -> str:
    """
    小学生にもわかる説明：
      Excel の見出し文字から前後の余計な空白を取り除き、
      同じ名前どうしを比べやすく整えます。
    """
    if value is None:
        return ""
    text = str(value)
    return text.strip()


def normalize_form_keys(data: Dict[str, str]) -> Dict[str, str]:
    """
    小学生にもわかる説明：
      フォームの項目名を Excel と同じように整えて、
      同じ列へ正しく書き込めるようにします。
    """
    # 初学者向け説明：正規化した結果を入れておく空の辞書を用意します。
    normalized: Dict[str, str] = {}
    # 初学者向け説明：フォームの項目名と値を順番に取り出して整えます。
    for key, value in data.items():
        # 初学者向け説明：余計な空白を取り除いて列名をそろえます。
        normalized_key = normalize_header_name(key)
        if not normalized_key:
            continue
        # 初学者向け説明：同じ列名が重複しても最初の値だけを残します。
        if normalized_key not in normalized:
            normalized[normalized_key] = value
    return normalized


def _normalize_cell_text(value: Any) -> str:
    """
    小学生にもわかる説明：
      Excel の値を文字として比べやすい形にそろえます。
    """
    # 初学者向け説明：空欄や None が来たときは比較しやすいよう空文字にします。
    if value is None:
        return ""

    # 初学者向け説明：数値の場合は 4.0 → 4 のように余計な小数を取り除きます。
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    # 初学者向け説明：それ以外は文字列化して返します。
    return str(value)


def _build_header_map_from_sheet(ws) -> Dict[str, int]:
    """
    小学生にもわかる説明：
      1行目の見出しを調べて、列名と列番号の対応表を作ります。
    """
    # 初学者向け説明：結果を入れる空の辞書を用意します。
    header_map: Dict[str, int] = {}

    # 初学者向け説明：1列ずつ見出しを確認し、余計な空白を取り除いた名前で登録します。
    for c in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=c).value
        if value is None:
            continue
        header_text = normalize_header_name(value)
        if not header_text or header_text in header_map:
            continue
        header_map[header_text] = c

    return header_map


def to_full_width(num: int) -> str:
    """
    小学生にもわかる説明：
      ふつうの数字(半角)を、見た目が広い数字(全角)に変えて返します。
    """
    return str(num).translate(_FW_TABLE)

# Windows の IME を制御するための準備（他OSでは使いません）
_IS_WINDOWS = sys.platform.startswith("win")
if _IS_WINDOWS:
    import ctypes
    _imm32 = ctypes.windll.imm32  # ImmAssociateContext を使う
    HWND = ctypes.wintypes.HWND if hasattr(
        ctypes, "wintypes") else ctypes.c_void_p
    HIMC = ctypes.c_void_p
    _imm32.ImmAssociateContext.argtypes = [HWND, HIMC]
    _imm32.ImmAssociateContext.restype = HIMC


# =========================================
# コンボボックスを左寄せ＆はみ出し防止にする仕組み
# =========================================
class LeftAlignDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        # 小学生にもわかる説明：
        #   リストの文字を左にそろえ、長すぎると「…」にします。
        super().initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter
        option.textElideMode = QtCore.Qt.ElideRight


def setup_left_aligned_combo(combo: QComboBox) -> None:
    """
    小学生にもわかる説明：
      この関数はプルダウンを『左寄せ』に直し、
      リストが狭くて文字がはみ出すのを防ぎます。
    """
    originally_editable = combo.isEditable()

    view = QListView(combo)
    view.setUniformItemSizes(True)
    view.setTextElideMode(QtCore.Qt.ElideRight)
    view.setItemDelegate(LeftAlignDelegate(view))
    combo.setView(view)

    combo.setLayoutDirection(QtCore.Qt.LeftToRight)

    if not originally_editable:
        combo.setEditable(True)
        combo.lineEdit().setReadOnly(True)
    if combo.lineEdit() is not None:
        combo.lineEdit().setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)

    try:
        width_hint = max(view.sizeHintForColumn(
            0), combo.view().sizeHintForColumn(0)) + 32
        view.setMinimumWidth(width_hint)
    except Exception:
        pass
# =========================================


# 数字入力専用のラインエディットです。フォーカス中は IME を完全に効かなくします。
class NumericLineEdit(QLineEdit):
    """
    小学生にもわかる説明：
      この入力欄は数字用です。カーソルが入っている間は
      日本語入力(IME)を使えないようにします。
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._prev_hints: Optional[QtCore.Qt.InputMethodHints] = None
        self._prev_ime_enabled: Optional[bool] = None
        # Windows 専用：元の IME ハンドルを覚えておく場所
        self._prev_himc = None

    def _hwnd(self) -> Optional[int]:
        if not _IS_WINDOWS:
            return None
        wid = self.winId()
        try:
            # PySide6 の winId() は sip.voidptr → int へ
            return int(wid)
        except Exception:
            return None

    def focusInEvent(self, event: QtGui.QFocusEvent) -> None:
        # いまのヒントと IME 状態を記録します
        self._prev_hints = self.inputMethodHints()
        self._prev_ime_enabled = self.testAttribute(
            QtCore.Qt.WA_InputMethodEnabled)

        # 半角英数字優先（数値入力向け）ヒント
        self.setInputMethodHints(
            self._prev_hints | QtCore.Qt.ImhLatinOnly | QtCore.Qt.ImhPreferNumbers)

        # Qt 側のIMEを無効（保険）
        self.setAttribute(QtCore.Qt.WA_InputMethodEnabled, False)

        # Windows では OS 側の IME をこのウィジェットから切り離します
        if _IS_WINDOWS:
            hwnd = self._hwnd()
            if hwnd:
                # 0 を関連付けると IME 無効化。戻り値が元の IME ハンドル
                self._prev_himc = _imm32.ImmAssociateContext(hwnd, HIMC(0))

        super().focusInEvent(event)

    def focusOutEvent(self, event: QtGui.QFocusEvent) -> None:
        # フォーカスが外れたら元の設定に戻します
        if self._prev_hints is not None:
            self.setInputMethodHints(self._prev_hints)
        if self._prev_ime_enabled is not None:
            self.setAttribute(QtCore.Qt.WA_InputMethodEnabled,
                              self._prev_ime_enabled)

        if _IS_WINDOWS:
            hwnd = self._hwnd()
            if hwnd:
                # 元の IME を戻します（他の欄で使えるように）
                _imm32.ImmAssociateContext(hwnd, self._prev_himc or HIMC(0))
                self._prev_himc = None

        super().focusOutEvent(event)

    def inputMethodEvent(self, event: QtGui.QInputMethodEvent) -> None:
        # 念のため：IME からの合成入力は無視（保険）
        event.ignore()


# === 起動時のエクセル読み込み待機ウインドウ ===
class LoadingSpinner(QtWidgets.QDialog):
    def __init__(self, parent: Optional[QtWidgets.QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowFlags(QtCore.Qt.Dialog | QtCore.Qt.FramelessWindowHint)
        self.setFixedSize(160, 160)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        self.indicator = _SpinnerWidget(self)
        layout.addWidget(self.indicator, alignment=QtCore.Qt.AlignCenter)

        label = QLabel("エクセルファイル\n読み込み中......", self)
        label.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(label, alignment=QtCore.Qt.AlignCenter)


class _SpinnerWidget(QtWidgets.QWidget):
    def __init__(self, parent: Optional[QtWidgets.QWidget] = None) -> None:
        super().__init__(parent)
        self._angle = 0
        self._pen_width = 8
        self._timer = QtCore.QTimer(self)
        self._timer.timeout.connect(self._rotate)
        self._timer.start(16)
        self.setFixedSize(64, 64)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)

    def _rotate(self) -> None:
        self._angle = (self._angle + 5) % 360
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent) -> None:
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.Antialiasing)
        rect = self.rect()
        painter.translate(rect.center())
        painter.rotate(self._angle)
        radius = min(rect.width(), rect.height()) / 2 - self._pen_width
        pen = QtGui.QPen(QtGui.QColor(0, 0, 255), self._pen_width)
        pen.setCapStyle(QtCore.Qt.RoundCap)
        painter.setPen(pen)
        painter.drawArc(QtCore.QRectF(-radius, -radius,
                        radius * 2, radius * 2), 0, 270 * 16)


# === シリンダー入力ユニット ===
class CylinderUnit(QWidget):
    def __init__(
        self,
        get_item_no: Callable[[], str],
        get_candidates: Callable[[str], List[str]],
    ) -> None:
        """シリンダー情報を1行分まとめる部品です。"""
        super().__init__()
        # 品目番号を取得する関数を保持します
        self._get_item_no = get_item_no
        # シリンダー候補を取得する関数を保持します
        self._get_candidates = get_candidates

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        row = QHBoxLayout()

        # 〇色目
        self.order_combo = QComboBox()
        self.order_combo.addItems(["0", "1"])
        self.order_combo.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToContents)
        row.addWidget(self.order_combo)

        # シリンダー番号（編集可）
        self.cylinder_combo = QComboBox()
        self.cylinder_combo.setEditable(True)
        self.cylinder_combo.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToContents)
        # 9桁の数字のみ入力できるようにします
        regex = QtCore.QRegularExpression(r"\d{9}")
        validator = QRegularExpressionValidator(regex)
        line_edit = self.cylinder_combo.lineEdit()
        if line_edit is not None:
            line_edit.setValidator(validator)
            line_edit.setInputMethodHints(QtCore.Qt.ImhDigitsOnly)
        # Excel から読み取った候補を表示します
        self.refresh_cylinder_list()
        row.addWidget(self.cylinder_combo)

        # 色名
        self.color_edit = QLineEdit()
        row.addWidget(self.color_edit)

        # ベタ巾（数値）
        self.width_edit = NumericLineEdit()
        dv = QDoubleValidator(0.0, 1e12, 3)
        dv.setNotation(QDoubleValidator.Notation.StandardNotation)
        dv.setLocale(QtCore.QLocale("C"))
        self.width_edit.setValidator(dv)
        self.width_edit.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)
        row.addWidget(self.width_edit)

        # 左寄せ＆はみ出し防止を適用
        setup_left_aligned_combo(self.order_combo)
        setup_left_aligned_combo(self.cylinder_combo)

        layout.addLayout(row)

    def refresh_cylinder_list(self) -> None:
        """
        Excel シートから取得したシリンダー候補を表示し直します。
        起動時に読み込んだ候補がすでに存在する場合は、無駄な再設定を避けます。
        """
        # すでに候補が設定されている場合は、これ以上の処理を行いません
        if self.cylinder_combo.count() > 0:
            return

        # いったんリストを空にします（初回のみ実行されます）
        self.cylinder_combo.clear()
        # 現在入力されている品目番号を取得します（候補自体は品目番号に依存しません）
        item_no = self._get_item_no()
        # すべての候補を取得してプルダウンに追加します
        candidates = self._get_candidates(item_no)
        self.cylinder_combo.addItems(candidates)



# === Excel の読み書き関数（省略なし・そのまま） ===
def _close_excel_workbook_if_open(path: str) -> None:
    """指定された Excel ファイルが開いていれば保存して閉じます"""
    # 初学者向け説明：最初にログへ記録し、どのファイルを対象にするか残します。
    LOGGER.info("_close_excel_workbook_if_open: 対象パス=%s", path)

    # 初学者向け説明：Windows だけが Excel を自動操作できるので対象 OS かどうか確認します。
    if not _IS_WINDOWS:
        # 初学者向け説明：Windows 以外では処理できないことを知らせます。
        LOGGER.info("_close_excel_workbook_if_open: Windows 以外のため処理を終了します")
        return

    # 初学者向け説明：win32com.client が準備できない場合は自動操作を諦めます。
    spec = importlib.util.find_spec("win32com.client")
    if spec is None:
        # 初学者向け説明：ライブラリが無いので終了したことを残します。
        LOGGER.info("_close_excel_workbook_if_open: win32com が見つからないため処理を終了します")
        return

    # 初学者向け説明：実際に win32com.client を読み込み、動作中の Excel を探します。
    win32_client = importlib.import_module("win32com.client")
    try:
        excel = win32_client.GetActiveObject("Excel.Application")
    except Exception:
        return

    # 初学者向け説明：比較を簡単にするため、パスを正規化します。
    target = os.path.normcase(os.path.normpath(os.path.abspath(path)))

    try:
        workbooks = excel.Workbooks
        count = workbooks.Count
    except Exception:
        return

    # 初学者向け説明：開いているブックを後ろから調べ、目的のファイルがあれば保存して閉じます。
    for idx in range(count, 0, -1):
        try:
            workbook = workbooks.Item(idx)
        except Exception:
            continue

        try:
            current = os.path.normcase(os.path.normpath(str(workbook.FullName)))
        except Exception:
            continue

        if current != target:
            continue

        try:
            workbook.Save()
            # 初学者向け説明：保存を指示したことをログに書きます。
            LOGGER.info("_close_excel_workbook_if_open: Excel へ保存命令を発行しました")
        except Exception:
            pass

        try:
            workbook.Close(SaveChanges=False)
            # 初学者向け説明：保存せずに閉じる命令を出したことを残します。
            LOGGER.info("_close_excel_workbook_if_open: Excel ブックを閉じました")
        except Exception:
            try:
                workbook.Close()
            except Exception:
                pass
        break

    # 初学者向け説明：最後に処理が終わったことを知らせます。
    LOGGER.info("_close_excel_workbook_if_open: 処理を完了しました")


def _try_upsert_with_excel(path: str, normalized_data: Dict[str, str], sheet_name: str,
                           save_mode: str) -> Optional[int]:
    """可能なら Excel 本体を使って安全に保存します。"""
    # 初学者向け説明：処理の開始をログへ残し、条件確認の準備をします。
    LOGGER.info(
        "_try_upsert_with_excel: 開始 path=%s sheet=%s mode=%s", path, sheet_name, save_mode
    )

    # 初学者向け説明：Windows 以外では Excel を直接操作できないので処理を打ち切ります。
    if not _IS_WINDOWS:
        LOGGER.info("_try_upsert_with_excel: Windows 以外のため処理を終了します")
        return None

    # 初学者向け説明：win32com.client を見つけられなければ Excel の自動操作はできません。
    spec = importlib.util.find_spec("win32com.client")
    if spec is None:
        LOGGER.info("_try_upsert_with_excel: win32com が見つからないため処理を終了します")
        return None

    # 初学者向け説明：実際に win32com.client を読み込み、Excel を起動（または接続）します。
    win32_client = importlib.import_module("win32com.client")
    excel = None
    workbook = None
    # 初学者向け説明：後で元に戻すため、「上書き確認の警告設定」の今の値を控える場所を用意します。
    previous_alert_before_overwriting = None
    alert_before_overwriting_supported = False
    try:
        try:
            excel = win32_client.DispatchEx("Excel.Application")
            LOGGER.info("_try_upsert_with_excel: Excel.Application を DispatchEx で取得しました")
        except Exception:
            try:
                excel = win32_client.Dispatch("Excel.Application")
                LOGGER.info("_try_upsert_with_excel: 既存の Excel.Application へ接続しました")
            except Exception:
                LOGGER.info("_try_upsert_with_excel: Excel を取得できず処理を終了します")
                return None

        # 初学者向け説明：画面に表示せず、警告も出さないように設定します。
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            # 初学者向け説明：上書き前に出る確認の設定を記録し、上書き警告が出ないように一時的に無効化します。
            previous_alert_before_overwriting = excel.AlertBeforeOverwriting
            excel.AlertBeforeOverwriting = False
            alert_before_overwriting_supported = True
        except Exception:
            # 初学者向け説明：古い Excel などで設定できない場合は、無理に触らずそのまま進めます。
            alert_before_overwriting_supported = False

        try:
            # 初学者にもわかる説明：Excel が「読み取り専用で開きますか？」と聞いてこないように、
            #   あらかじめ編集できる設定でブックを開くよう命令します。
            workbook = excel.Workbooks.Open(
                path,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
            )
            LOGGER.info("_try_upsert_with_excel: ブックを編集モードで開きました")
        except Exception:
            LOGGER.info("_try_upsert_with_excel: ブックを開けず処理を終了します")
            return None

        try:
            worksheet = workbook.Worksheets(sheet_name)
            LOGGER.info("_try_upsert_with_excel: 対象シートを取得しました")
        except Exception as exc:
            raise ValueError(f"シート『{sheet_name}』がありません") from exc

        # 初学者向け説明：見出し行（1行目）を調べ、列名から列番号の辞書を作ります。
        header_map: Dict[str, int] = {}
        xl_to_left = -4159  # Excel 定数：左方向へ移動（xlToLeft）
        try:
            last_header_col = int(worksheet.Cells(1, worksheet.Columns.Count)
                                  .End(xl_to_left).Column)
        except Exception:
            last_header_col = 0
        if last_header_col < 0:
            last_header_col = 0
        for col in range(1, last_header_col + 1):
            header_value = worksheet.Cells(1, col).Value
            if header_value is None:
                continue
            # 初学者向け説明：列名を文字に直し、余計な空白を取り除いて扱いやすくします。
            header_text = normalize_header_name(header_value)
            if not header_text or header_text in header_map:
                continue
            header_map[header_text] = col

        LOGGER.info(
            "_try_upsert_with_excel: 見出しを確認しました headers=%s",
            list(header_map.keys()),
        )

        # 初学者向け説明：必要な列が無ければ保存できないのでエラーにします。
        if "品目番号" not in header_map:
            raise ValueError("『品目番号』の列が見つかりません")

        item_column = header_map["品目番号"]
        xl_up = -4162  # Excel 定数：上方向へ移動（xlUp）
        try:
            last_filled_row = int(worksheet.Cells(worksheet.Rows.Count, item_column)
                                  .End(xl_up).Row)
        except Exception:
            last_filled_row = 1
        if last_filled_row < 1:
            last_filled_row = 1

        LOGGER.info(
            "_try_upsert_with_excel: 品目番号の最終行を確認しました last_row=%d",
            last_filled_row,
        )

        # 初学者向け説明：上書き保存の場合は既存行を探し、新規の場合は最後尾に追加します。
        target_row: Optional[int] = None
        item_value = normalized_data.get("品目番号", "")
        if save_mode == "上書き保存":
            for row in range(2, last_filled_row + 1):
                cell_value = worksheet.Cells(row, item_column).Value
                if str(cell_value or "").strip() == item_value:
                    target_row = row
                    break
            if target_row is None:
                LOGGER.info("_try_upsert_with_excel: 上書き対象が見つからずエラーにします")
                raise ValueError("上書き対象の行が見つかりません。先に該当データを読み込んでください。")
            LOGGER.info(
                "_try_upsert_with_excel: 上書き対象行を特定しました row=%s item=%s",
                target_row,
                item_value,
            )
        else:
            target_row = last_filled_row + 1
            LOGGER.info(
                "_try_upsert_with_excel: 新規登録の書き込み行を決定しました row=%s",
                target_row,
            )

        # 初学者向け説明：列名ごとに対応する値を取り出して書き込みます。
        for header_key, column in header_map.items():
            if header_key not in normalized_data:
                continue
            value = normalized_data.get(header_key)
            worksheet.Cells(target_row, column).Value = "" if value is None else str(value)

        LOGGER.info("_try_upsert_with_excel: Excel シートへ書き込みました row=%s", target_row)

        # 初学者向け説明：ここまでの変更を Excel に保存します。
        workbook.Save()
        LOGGER.info("_try_upsert_with_excel: Excel ブックを保存しました")
        return target_row

    except ValueError:
        # 初学者向け説明：入力不足など明確なエラーはそのまま呼び出し元へ伝えます。
        raise
    except Exception:
        # 初学者向け説明：想定外のエラーは Excel 保存に失敗したとして、後段の処理へ委ねます。
        LOGGER.exception("_try_upsert_with_excel: 想定外のエラーが発生しました")
        return None
    finally:
        # 初学者向け説明：Excel を起動したままにしないよう、必ず後片付けします。
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                if alert_before_overwriting_supported:
                    try:
                        # 初学者向け説明：作業前に覚えておいた上書き警告の設定を元通りに戻します。
                        excel.AlertBeforeOverwriting = previous_alert_before_overwriting
                    except Exception:
                        pass
                excel.Quit()
                LOGGER.info("_try_upsert_with_excel: Excel アプリケーションを終了しました")
            except Exception:
                pass


def _verify_saved_row(path: str, sheet_name: str, target_row: int,
                      normalized_data: Dict[str, str]) -> bool:
    """
    小学生にもわかる説明：
      Excel に書き込んだ内容が本当に反映されたかを読み直して確認します。
    """
    # 初学者向け説明：念のため失敗を前提にしておき、途中で問題があれば False を返します。
    wb = None
    try:
        # 初学者向け説明：openpyxl で Excel ファイルを開きます。
        wb = load_workbook(path, keep_vba=True, data_only=False)
    except Exception:
        LOGGER.exception("_verify_saved_row: ブックを開けませんでした")
        return False

    try:
        if sheet_name not in wb.sheetnames:
            LOGGER.error("_verify_saved_row: シート %s が見つかりません", sheet_name)
            return False

        ws = wb[sheet_name]

        # 初学者向け説明：列の対応表を作り、存在しない列は確認対象から外します。
        header_map = _build_header_map_from_sheet(ws)
        if not header_map:
            LOGGER.error("_verify_saved_row: 見出しを取得できませんでした")
            return False

        if target_row < 1 or target_row > ws.max_row:
            LOGGER.warning(
                "_verify_saved_row: 指定行が範囲外のため確認できません row=%s max=%s",
                target_row,
                ws.max_row,
            )
            return False

        for header_key, expected_value in normalized_data.items():
            column = header_map.get(header_key)
            if column is None:
                continue

            expected_text = _normalize_cell_text(expected_value)
            actual_cell = ws.cell(row=target_row, column=column).value
            actual_text = _normalize_cell_text(actual_cell)

            if actual_text != expected_text:
                LOGGER.warning(
                    "_verify_saved_row: 値が一致しません column=%s expected=%s actual=%s row=%s",
                    header_key,
                    expected_text,
                    actual_text,
                    target_row,
                )
                return False

        # 初学者向け説明：すべて一致したので True を返します。
        return True
    finally:
        if wb is not None:
            wb.close()


def _upsert_with_openpyxl(path: str, normalized_data: Dict[str, str], sheet_name: str,
                          save_mode: str, target_row_hint: Optional[int] = None) -> int:
    """
    小学生にもわかる説明：
      openpyxl を使って Excel ファイルへ書き込みます。
    """
    try:
        # 初学者向け説明：Excel を誰かが開いていても読めるよう、例外を補足して案内します。
        wb = load_workbook(path, keep_vba=True, data_only=False)
    except PermissionError as e:
        LOGGER.exception("_upsert_with_openpyxl: ファイルを開けませんでした")
        raise PermissionError(
            "Excel が開いている可能性があります。Excel を手動で閉じてから再度お試しください。"
        ) from e

    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート『{sheet_name}』がありません")

        ws = wb[sheet_name]

        header_map = _build_header_map_from_sheet(ws)
        if "品目番号" not in header_map:
            raise ValueError("『品目番号』の列が見つかりません")

        item_column = header_map["品目番号"]
        target_row: Optional[int] = None

        if save_mode == "上書き保存":
            item_value = normalized_data.get("品目番号", "")

            # 初学者向け説明：Excel COM が教えてくれた行番号があれば先に照合します。
            if target_row_hint is not None and target_row_hint >= 2:
                cell_value = ws.cell(row=target_row_hint, column=item_column).value
                if _normalize_cell_text(cell_value) == item_value:
                    target_row = target_row_hint
                    LOGGER.info(
                        "_upsert_with_openpyxl: COM が返した行番号を再利用します row=%s",
                        target_row,
                    )
                else:
                    LOGGER.info(
                        "_upsert_with_openpyxl: COM の行番号と値が一致しません row=%s",
                        target_row_hint,
                    )

            # 初学者向け説明：ヒントが使えない場合は上から順番に探します。
            if target_row is None:
                for r in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=r, column=item_column).value
                    if _normalize_cell_text(cell_value) == item_value:
                        target_row = r
                        break

            if target_row is None:
                LOGGER.info("upsert_record_to_xlsm: 上書き対象が見つからずエラーを投げます")
                raise ValueError("上書き対象の行が見つかりません。先に該当データを読み込んでください。")

            LOGGER.info("upsert_record_to_xlsm: 上書き対象行 row=%s", target_row)
        else:
            # 初学者向け説明：新規登録時はヒントがあればその行に、無ければ末尾に追加します。
            if target_row_hint is not None and target_row_hint >= 2:
                target_row = target_row_hint
                LOGGER.info(
                    "_upsert_with_openpyxl: COM が案内した新規行を使用します row=%s",
                    target_row,
                )
            else:
                last_filled_row = 1
                for r in range(ws.max_row, 1, -1):
                    value = ws.cell(row=r, column=item_column).value
                    if value not in (None, ""):
                        last_filled_row = r
                        break
                target_row = last_filled_row + 1
                LOGGER.info("upsert_record_to_xlsm: 新規登録の書き込み行 row=%s", target_row)

        # 初学者向け説明：対象となる列だけを順番に上書きします。
        for header_key, column in header_map.items():
            if header_key not in normalized_data:
                continue
            value = normalized_data.get(header_key)
            ws.cell(row=target_row, column=column).value = "" if value is None else str(value)

        wb.save(path)
        return target_row
    finally:
        wb.close()


def read_record_from_xlsm(path: str, item_no: str, sheet_name: str) -> Optional[Dict[str, str]]:
    wb = load_workbook(path, keep_vba=True, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート『{sheet_name}』がありません")
    ws = wb[sheet_name]

    # 初学者向け説明：共通の関数を使って列名と列番号の対応表を作ります。
    header_map = _build_header_map_from_sheet(ws)

    if "品目番号" not in header_map:
        raise ValueError("『品目番号』の列が見つかりません")

    col_item = header_map["品目番号"]
    target_row = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=col_item).value or "") == item_no:
            target_row = r
            break
    if target_row is None:
        return None

    return {name: str(ws.cell(row=target_row, column=header_map.get(name)).value or "")
            for name in header_map.keys()}


def upsert_record_to_xlsm(path: str, data: Dict[str, str], sheet_name: str, save_mode: str) -> None:
    # 初学者向け説明：保存作業の前に Excel が開いていれば自動で閉じて安全にします。
    LOGGER.info(
        "upsert_record_to_xlsm: 処理開始 path=%s sheet=%s mode=%s",
        path,
        sheet_name,
        save_mode,
    )

    # 初学者向け説明：列名を揃えておき、あとで Excel と確実に突き合わせられるようにします。
    normalized_data = normalize_form_keys(data)
    LOGGER.info(
        "upsert_record_to_xlsm: 入力データを正規化しました data=%s",
        _summarize_for_log(normalized_data),
    )

    _close_excel_workbook_if_open(path)

    # 初学者向け説明：まずは Excel 本体での保存を試み、成功した行番号を受け取ります。
    target_row = _try_upsert_with_excel(path, normalized_data, sheet_name, save_mode)
    if target_row is not None:
        # 初学者向け説明：念のため openpyxl で実データを読み直し、書き込み結果を確認します。
        if _verify_saved_row(path, sheet_name, target_row, normalized_data):
            LOGGER.info(
                "upsert_record_to_xlsm: Excel COM 保存が成功しました row=%s",
                target_row,
            )
            return
        LOGGER.warning(
            "upsert_record_to_xlsm: Excel COM の保存内容を確認できなかったため再保存します"
        )

    LOGGER.info("upsert_record_to_xlsm: openpyxl による保存へ切り替えます")

    # 初学者向け説明：Excel COM で使用した行番号をヒントとして再利用しつつ保存します。
    target_row = _upsert_with_openpyxl(
        path,
        normalized_data,
        sheet_name,
        save_mode,
        target_row_hint=target_row,
    )

    LOGGER.info("upsert_record_to_xlsm: openpyxl で保存を完了しました row=%s", target_row)


# === 起動時データ抽出関数（省略なし・そのまま） ===
def extract_initial_data(path: str, progress: Optional[Callable[[int, int], None]] = None) -> Dict[str, List[List[Any]]]:
    wb = load_workbook(path, keep_vba=True, data_only=False)
    result: Dict[str, List[List[Any]]] = {}
    sheets = ("受注データ", "シリンダーデータ")
    total = len(sheets)
    for idx, sheet in enumerate(sheets, start=1):
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            result[sheet] = _extract_range_from_sheet(ws)
        if progress is not None:
            progress(idx, total)
    return result


def _extract_range_from_sheet(ws) -> List[List[Any]]:
    max_col = ws.max_column
    while max_col > 0 and ws.cell(row=1, column=max_col).value in (None, ""):
        max_col -= 1

    header_map: Dict[str, int] = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        # 初学者向け説明：見出しの文字から余計な空白を取り除いて、列名をそろえています。
        header_text = normalize_header_name(v)
        if not header_text or header_text in header_map:
            continue
        header_map[header_text] = c

    last_row = 1
    for key in ("品目番号", "品目番号+刷順"):
        col = header_map.get(key)
        if col is None:
            continue
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=col).value not in (None, ""):
                if r > last_row:
                    last_row = r
                break

    data: List[List[Any]] = []
    for r in range(1, last_row + 1):
        row_values: List[Any] = []
        for c in range(1, max_col + 1):
            row_values.append(ws.cell(row=r, column=c).value)
        data.append(row_values)
    return data

# === 辞書作成と検索の処理例 ===
# 辞書を作成する処理です。品目番号をキーにします。
# record_index = {row_key: row_data for row_key, row_data in ...}
#
# 辞書から該当レコードを取得します。存在しない場合は None を返します。
# result = record_index.get(search_value)
# ===================================

def build_record_index(data: List[List[Any]], columns: List[str]) -> Dict[str, Dict[str, Dict[str, str]]]:
    """
    小学生にもわかる説明：
      表の1行目を見出しにして、指定された列の値をキーにした辞書を作ります。
      この辞書を使うと、同じ列を1行ずつ調べるより速く探せます。
    """
    # 返り値の入れ物を用意します
    index: Dict[str, Dict[str, Dict[str, str]]] = {}
    if not data:
        return index

    # 1行目は見出しなので、空白を取り除いた名前にそろえておきます
    headers = [normalize_header_name(v) for v in data[0]]

    for column in columns:
        # 欲しい列が無い場合は飛ばします
        if column not in headers:
            continue
        idx = headers.index(column)
        col_dict: Dict[str, Dict[str, str]] = {}
        for row in data[1:]:
            # 行から品目番号などのキーを取り出します
            if idx < len(row) and row[idx] is not None:
                key = str(row[idx])
                # 行全体を {列名: 値} の辞書に変換します
                col_dict[key] = {h: (str(row[i]) if i < len(row) and row[i] is not None else "")
                                 for i, h in enumerate(headers) if h}
        index[column] = col_dict
    return index


def find_record_by_column(index_map: Dict[str, Dict[str, Dict[str, str]]], column: str, value: str) -> Optional[Dict[str, str]]:
    """
    小学生にもわかる説明：
      作っておいた辞書から、ほしい値の行をすぐに取り出します。
      その値が無ければ、何も見つからなかった印として None を返します。
    """
    start_time = time.perf_counter()  # 処理開始時刻を記録します
    # 指定された列の辞書を取り出します
    column_dict = index_map.get(column, {})
    # 辞書から直接レコードを取り出します。存在しなければ None です
    record = column_dict.get(value)
    elapsed = time.perf_counter() - start_time  # 処理にかかった時間を計算します
    print(f"[LOG] find_record_by_column: {elapsed:.3f} 秒")  # かかった時間を表示します
    return record


# === Excel 検索を別スレッドで行うワーカー ===
class FetchWorker(QtCore.QThread):
    """
    小学生にもわかる説明：
      このクラスは時間のかかる検索を別のスレッドで行います。
      これにより画面の操作が止まらないようにします。
    """

    # 検索結果をメインスレッドへ渡すためのシグナルです。
    finished = QtCore.Signal(dict)

    def __init__(self, index_map: Dict[str, Dict[str, Dict[str, str]]], item_no: str) -> None:
        """検索に必要な情報を受け取り、保存します"""
        super().__init__()
        # 検索に利用する辞書を覚えておきます
        self._index_map = index_map
        # 品目番号も覚えておきます
        self._item_no = item_no

    def run(self) -> None:
        """別スレッドで実際の検索を行います"""
        try:
            # 重い検索処理をここで行います
            record = find_record_by_column(self._index_map, "品目番号", self._item_no)
            # 検索が終わったらシグナルで結果を返します
            self.finished.emit({"record": record, "item_no": self._item_no})
        except Exception as e:
            # エラーが出た場合はエラーメッセージを渡します
            self.finished.emit({"record": None, "item_no": self._item_no, "error": str(e)})


# === マテリアル風のカード ===
class Card(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")
        self.setFrameShape(QFrame.Shape.NoFrame)
        self.setStyleSheet("""
            QFrame#card {
                background: #FFFFFF;
                border-radius: 12px;
                border: 1px solid rgba(0,0,0,0.08);
            }
        """)


# === メイン画面 ===
class MainWindow(QMainWindow):
    def __init__(self, layout_path: str):
        super().__init__()

        self.config = self._load_layout(layout_path)
        self.default_font_size = int(self.config.get("font_size", 12))

        win = self.config.get("window", {})
        self.setWindowTitle(win.get("title", "フォーム"))
        self.resize(int(win.get("width", 980)), int(win.get("height", 680)))

        self.excel_sheet = self.config.get("excel", {}).get("sheet", "受注データ")
        self.path_store = os.path.join(
            os.path.dirname(layout_path), "data_file_path.txt")
        self.current_xlsm: Optional[str] = self.load_xlsm_path()

        self.preloaded_data: Dict[str, List[List[Any]]] = {}
        # 検索を速くするための辞書を保存する入れ物です
        self.record_index: Dict[str, Dict[str, Dict[str, Dict[str, str]]]] = {}
        # 起動時に計算したシリンダー番号の候補を保存する入れ物です
        self._cylinder_candidates_cache: List[str] = []
        while self.current_xlsm is not None:
            spinner = LoadingSpinner(self)
            spinner.show()
            QApplication.processEvents()

            container = {"data": {}, "error": None}

            def load() -> None:
                try:
                    container["data"] = extract_initial_data(self.current_xlsm)
                except Exception as e:
                    container["error"] = e

            thread = threading.Thread(target=load)
            thread.start()

            while thread.is_alive():
                QApplication.processEvents()
                QtCore.QThread.msleep(50)

            thread.join()
            spinner.close()

            if container["error"] is not None:
                QMessageBox.warning(
                    self,
                    "読み込みエラー",
                    f"初期データの読み込みに失敗しました: {container['error']}",
                )
                new_path = self.ask_xlsm_path()
                if new_path:
                    with open(self.path_store, "w", encoding="utf-8") as f:
                        f.write(new_path)
                    self.current_xlsm = new_path
                    continue
                self.current_xlsm = None
            else:
                self.preloaded_data = container["data"]
                # 辞書を作成する処理です。品目番号をキーにします。
                self.record_index = {
                    sheet: build_record_index(rows, SEARCH_COLUMNS)
                    for sheet, rows in self.preloaded_data.items()
                }
                # シリンダー番号の候補を起動時に一度だけ計算して保存します
                self._cylinder_candidates_cache = self._collect_cylinder_candidates()
                break

        if self.current_xlsm is None:
            QMessageBox.information(
                self,
                "設定情報",
                "データファイルが設定されていないため、読み書きは行えません。",
            )

        self.status = QStatusBar(self)
        self.setStatusBar(self.status)

        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)

        container = QWidget()
        root = QVBoxLayout(container)
        root.setContentsMargins(24, 24, 24, 24)
        root.setSpacing(24)

        if "title" in self.config:
            title = QLabel(self.config["title"])
            title.setStyleSheet("font-size:22px; font-weight:600;")
            root.addWidget(title)

        self.card = Card()
        wrap = QVBoxLayout(self.card)
        wrap.setContentsMargins(24, 24, 24, 24)
        wrap.setSpacing(16)

        self.grid = QGridLayout()
        self.grid.setHorizontalSpacing(16)
        self.grid.setVerticalSpacing(12)
        ncols = int(self.config.get("grid_columns", 4))
        for i in range(ncols * 2):
            stretch = 1 if i % 2 == 1 else 0
            self.grid.setColumnStretch(i, stretch)

        self.widgets: Dict[str, QtWidgets.QWidget] = {}
        self.save_button: Optional[QPushButton] = None
        self._build_from_config(self.config.get("fields", []), ncols)

        self._last_fetched_item: str = ""

        item_widget = self.widgets.get("品目番号")
        if isinstance(item_widget, QLineEdit):
            item_widget.textChanged.connect(self.on_item_no_changed)
        self.update_button_states()

        color_widget = self.widgets.get("色数")
        if isinstance(color_widget, QLineEdit):
            color_widget.textChanged.connect(self.on_color_count_changed)

        self.cyl_title = QLabel("シリンダーデータ登録")
        self.cyl_title.setStyleSheet("font-weight:600;")

        self.cylinder_layout = QVBoxLayout()
        self.cylinder_layout.setSpacing(8)

        # 入力欄の見出しを横一列に並べるためのレイアウトを準備します
        self.cylinder_header = QHBoxLayout()
        header_labels = ["〇色目", "シリンダー番号", "色名", "ベタ巾"]
        for text in header_labels:
            lbl = QLabel(text)
            lbl.setStyleSheet("font-weight:600;")
            self.cylinder_header.addWidget(lbl)
        self.cylinder_layout.addLayout(self.cylinder_header)

        # 入力欄を作成中であることを知らせるラベルを用意し、普段は隠しておきます
        self.cylinder_status_label = QLabel("")
        self.cylinder_status_label.hide()
        self.cylinder_layout.addWidget(self.cylinder_status_label)

        self.cylinder_units: List[CylinderUnit] = []
        # 先頭の色番号コンボがどれかを覚えて、不要な二重接続を避けます
        self._first_order_combo: Optional[QComboBox] = None

        wrap.addLayout(self.grid)
        wrap.addWidget(self.cyl_title)
        wrap.addLayout(self.cylinder_layout)
        root.addWidget(self.card)
        scroll.setWidget(container)
        self.setCentralWidget(scroll)

    def _load_layout(self, path: str) -> Dict[str, Any]:
        if not os.path.exists(path):
            raise FileNotFoundError(f"設計図ファイルが見つかりません: {path}")
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _build_from_config(self, fields: list, ncols: int):
        for f in fields:
            ftype = f.get("type", "")
            row = int(f.get("row", 0))
            col = int(f.get("col", 0))
            col_span = int(f.get("col_span", 1))
            grid_col_label = col * 2
            grid_col_edit = col * 2 + 1
            grid_span = max(1, min(ncols - col, col_span)) * 2 - 1

            font_size = int(f.get("font_size", self.default_font_size))

            if ftype == "header":
                lbl = QLabel(f.get("text", ""))
                lbl.setStyleSheet(
                    f"font-size:{font_size}px; font-weight:600; color:#333;")
                self.grid.addWidget(lbl, row, 0, 1, ncols * 2)
                continue

            if ftype in ("line", "text"):
                label_text = f.get("label", "")
                key = f.get("key", label_text)

                lbl = QLabel(label_text)
                lbl.setStyleSheet(f"color:#333; font-size:{font_size}px;")
                self.grid.addWidget(lbl, row, grid_col_label)

                if ftype == "text":
                    edit = QPlainTextEdit()
                    h = int(f.get("height", 120))
                    edit.setFixedHeight(h)
                else:
                    val = f.get("validator", "")
                    if val in ("int", "float"):
                        edit = NumericLineEdit()
                    else:
                        edit = QLineEdit()

                    if val == "int":
                        imin = int(f.get("min", 0))
                        imax = int(f.get("max", 2147483647))
                        iv = QIntValidator(imin, imax)
                        iv.setLocale(QtCore.QLocale("C"))
                        edit.setValidator(iv)
                        edit.setInputMethodHints(QtCore.Qt.ImhDigitsOnly)
                    elif val == "float":
                        fmin = float(f.get("min", 0.0))
                        fmax = float(f.get("max", 1e12))
                        dec = int(f.get("decimals", 3))
                        v = QDoubleValidator(fmin, fmax, dec)
                        v.setNotation(
                            QDoubleValidator.Notation.StandardNotation)
                        v.setLocale(QtCore.QLocale("C"))
                        edit.setValidator(v)
                        edit.setInputMethodHints(QtCore.Qt.ImhPreferNumbers)

                edit.setStyleSheet(f"""
                    QLineEdit {{
                        padding: 8px 10px;
                        border-radius: 8px;
                        border:1px solid rgba(0,0,0,0.15);
                        background:#FAFAFA;
                        font-size:{font_size}px;
                    }}
                    QPlainTextEdit {{
                        padding: 10px;
                        border-radius: 10px;
                        border:1px solid rgba(0,0,0,0.15);
                        background:#FAFAFA;
                        font-size:{font_size}px;
                    }}
                    QLineEdit:focus, QPlainTextEdit:focus {{
                        border:1.4px solid #2962FF;
                        background:#FFFFFF;
                    }}
                """)

                w = int(f.get("width", 0))
                if w > 0:
                    edit.setFixedWidth(w)
                self.grid.addWidget(edit, row, grid_col_edit, 1, grid_span)
                self.widgets[key] = edit
                continue

            if ftype == "button":
                text = f.get("text", "ボタン")
                action = f.get("action", "")
                btn = QPushButton(text)
                btn.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background:#2962FF;
                        color:white;
                        padding:8px 16px;
                        border:none;
                        border-radius:8px;
                        font-weight:600;
                        font-size:{font_size}px;
                    }}
                    QPushButton:hover {{ background:#2F6BFF; }}
                    QPushButton:pressed {{ background:#2554CC; }}
                    QPushButton:disabled {{
                        background:#E0E0E0;
                        color:#9E9E9E;
                    }}
                """)
                w = int(f.get("width", 0))
                if w > 0:
                    btn.setFixedWidth(w)
                self.grid.addWidget(btn, row, grid_col_edit,
                                    1, max(1, grid_span))
                if action == "save":
                    self.save_button = btn
                    btn.clicked.connect(self.on_save)
                elif action == "clear":
                    btn.clicked.connect(self.on_clear)
                elif action == "close":
                    btn.clicked.connect(self.close)
                continue

    def on_item_no_changed(self, text: str) -> None:
        start_time = time.perf_counter()  # 処理開始時刻を記録します
        """品目番号の入力が変わったときの共通処理です。"""
        self.update_button_states()

        # シリンダー番号の候補は起動時に読み込まれるため、
        # ここでの再読み込みは行いません

        item_no = text.strip()
        has_file = self.current_xlsm is not None
        is_item_eight_digits = re.fullmatch(r"\d{8}", item_no) is not None

        if has_file and is_item_eight_digits and item_no != self._last_fetched_item:
            self.on_fetch()
            self._last_fetched_item = item_no

        elapsed = time.perf_counter() - start_time  # 処理にかかった時間を計算します
        print(f"[LOG] on_item_no_changed: {elapsed:.3f} 秒")  # かかった時間を表示します

    def update_button_states(self) -> None:
        """入力内容に応じて『保存』ボタンの状態を切り替えます。"""
        item_widget = self.widgets.get("品目番号")
        item_text = ""
        if isinstance(item_widget, QLineEdit):
            item_text = item_widget.text().strip()

        has_file = self.current_xlsm is not None
        is_save_valid = bool(item_text) and has_file

        save_text = "新規登録"
        if item_text and has_file:
            sheet_index = self.record_index.get(self.excel_sheet)
            if sheet_index:
                try:
                    exists = find_record_by_column(
                        sheet_index, "品目番号", item_text)
                    if exists is not None:
                        save_text = "上書き保存"
                except Exception:
                    pass

        if self.save_button is not None:
            self.save_button.setEnabled(is_save_valid)
            self.save_button.setText(save_text)

    def on_color_count_changed(self, text: str) -> None:
        """色数に応じてシリンダー入力欄を増減させます。"""
        # 数字に変換できない入力は 0 色として扱います
        new_count = int(text) if text.isdigit() else 0
        current_count = len(self.cylinder_units)

        # 件数が変わっていないときは何もしません
        if new_count == current_count:
            return

        # 入力欄を作成・削除していることを画面に表示します
        self._set_generation_status(True)
        try:
            if new_count > current_count:
                # 増えた分だけ新しいシリンダー入力行を追加します
                for _ in range(new_count - current_count):
                    unit = CylinderUnit(self._get_item_no, self._get_cylinder_candidates)
                    self.cylinder_layout.addWidget(unit)
                    self.cylinder_units.append(unit)
            else:
                # 減った分だけ後ろから入力行を削除します
                self._remove_cylinder_units(current_count - new_count)

            # 色番号を 1 から順番に並べ、必要なら 0 も選べるようにします
            self.update_color_numbers(1)
        finally:
            # 先頭行のシグナル接続を整理し、完了後にメッセージを消します
            self._refresh_first_color_signal()
            self._set_generation_status(False)

    def _get_item_no(self) -> str:
        """現在入力されている品目番号を取得します。"""
        w = self.widgets.get("品目番号")
        if isinstance(w, QLineEdit):
            return w.text().strip()
        return ""

    def _collect_cylinder_candidates(self) -> List[str]:
        """
        小学生にもわかる説明：
          Excel の「シリンダーデータ」からシリンダー番号の候補を
          一度だけ集めて覚えておきます。
        """
        # 事前に読み込んだ「シリンダーデータ」シートを取り出します
        data = self.preloaded_data.get("シリンダーデータ")
        if not data:
            return []
        # 1 行目から列名を集めます
        headers = [str(v) if v is not None else "" for v in data[0]]
        # 「品目番号+刷順列」の列を探します。古い列名にも対応します
        cyl_header = "品目番号+刷順列"
        if cyl_header not in headers:
            if "品目番号+刷順" in headers:
                cyl_header = "品目番号+刷順"
            else:
                return []
        idx_cyl = headers.index(cyl_header)
        result: List[str] = []
        seen = set()
        # 2 行目以降を順番に調べて、9桁の数字だけを重複なく集めます
        for row in data[1:]:
            cyl_cell = (
                str(row[idx_cyl])
                if idx_cyl < len(row) and row[idx_cyl] is not None
                else ""
            )
            if re.fullmatch(r"\d{9}", cyl_cell) and cyl_cell not in seen:
                result.append(cyl_cell)
                seen.add(cyl_cell)
        return result

    def _get_cylinder_candidates(self, item_no: str) -> List[str]:
        """
        小学生にもわかる説明：
          起動時に集めたシリンダー番号の候補をそのまま返します。
          品目番号は受け取りますが計算には使いません。
        """
        # 起動時に計算しておいた候補を返します
        return self._cylinder_candidates_cache

    def _clear_cylinder_units(self) -> None:
        """シリンダー入力欄をすべて取り除きます。"""
        # いま残っている行数ぶんだけ削除し、リストも空にします
        self._remove_cylinder_units(len(self.cylinder_units))
        self._refresh_first_color_signal()

    def _remove_cylinder_units(self, count: int) -> None:
        """指定された数だけ末尾からシリンダー入力欄を削除します。"""
        # 削除数が現在の行数を超えないように調整します
        removal_count = min(count, len(self.cylinder_units))
        for _ in range(removal_count):
            unit = self.cylinder_units.pop()
            # レイアウトから取り外し、ウィジェットを破棄します
            self.cylinder_layout.removeWidget(unit)
            unit.deleteLater()

    def _refresh_first_color_signal(self) -> None:
        """先頭行のシグナル接続を整理して過剰な呼び出しを防ぎます。"""
        # 以前に覚えていた先頭のコンボボックスとの接続を外します
        if self._first_order_combo is not None:
            try:
                self._first_order_combo.currentTextChanged.disconnect(self.on_first_color_changed)
            except TypeError:
                pass
            self._first_order_combo = None

        # 新しく先頭が存在する場合だけ接続し直します
        if self.cylinder_units:
            first_combo = self.cylinder_units[0].order_combo
            first_combo.currentTextChanged.connect(self.on_first_color_changed)
            self._first_order_combo = first_combo

    def _set_generation_status(self, visible: bool) -> None:
        """入力欄生成中のメッセージ表示を切り替えます。"""
        if visible:
            # ラベルに文字を入れて表示し、すぐに画面へ反映させます
            self.cylinder_status_label.setText("入力欄生成中....")
            self.cylinder_status_label.show()
            QtWidgets.QApplication.processEvents()
        else:
            # 作業が終わったら文字を消して非表示に戻します
            self.cylinder_status_label.clear()
            self.cylinder_status_label.hide()

    def update_color_numbers(self, start: int) -> None:
        """表示される色番号を0または1から順番に並べ直します。"""

        # すべての行に対して、先頭から順番に色番号を設定します
        for idx, unit in enumerate(self.cylinder_units):
            unit.order_combo.blockSignals(True)  # 値を書き換える間はシグナルを止めます
            unit.order_combo.clear()             # 以前の選択肢を消します

            # 表示する番号を計算します（start が0なら0から、1なら1から）
            number = start + idx

            if idx == 0:
                # 先頭の行だけは0と1を選べるようにして、初期値を設定します
                unit.order_combo.addItems(["0", "1"])
                unit.order_combo.setCurrentText(str(number))
                unit.order_combo.setEnabled(True)
            else:
                # 2行目以降は計算した番号を表示し、変更できないようにします
                unit.order_combo.addItem(str(number))
                unit.order_combo.setCurrentIndex(0)
                unit.order_combo.setEnabled(False)

            unit.order_combo.blockSignals(False)  # シグナルを再び有効にします

    def on_first_color_changed(self, text: str) -> None:
        """先頭の色番号が0か1かで全体の番号を調整します。"""
        start = 0 if text.strip() == "0" else 1
        self.update_color_numbers(start)

    def collect_form_data(self) -> Dict[str, str]:
        # 初学者向け説明：画面の入力欄からテキストを集めて、Excelに渡す準備をします。
        d: Dict[str, str] = {}
        for k, w in self.widgets.items():
            if isinstance(w, QPlainTextEdit):
                d[k] = w.toPlainText().strip()
            elif isinstance(w, QLineEdit):
                d[k] = w.text().strip()

        # 初学者向け説明：シリンダー番号も専用の列へ入れられるように集めます。
        d.update(self._collect_cylinder_form_data())

        # 初学者向け説明：収集した内容をログに残し、後で保存時の状況を追えるようにします。
        LOGGER.info(
            "collect_form_data: 入力値を集計しました data=%s",
            _summarize_for_log(d),
        )
        return d

    def _collect_cylinder_form_data(self) -> Dict[str, str]:
        # 初学者向け説明：０色目から１０色目の列名を先に用意し、何も無ければ空欄にしておきます。
        cylinder_data: Dict[str, str] = {
            f"{to_full_width(order)}色目シリンダー": "" for order in range(0, 11)
        }

        # 初学者向け説明：各行の入力欄から番号を取り出し、対応する列へ入れます。
        for unit in self.cylinder_units:
            order_text = unit.order_combo.currentText().strip()
            if not order_text.isdigit():
                continue

            order = int(order_text)
            if 0 <= order <= 10:
                line = unit.cylinder_combo.lineEdit()
                value = ""
                if line is not None:
                    value = line.text().strip()
                cylinder_data[f"{to_full_width(order)}色目シリンダー"] = value

        # 初学者向け説明：列ごとの値をログへ記録し、どの番号にどんな値を入れたか確認できるようにします。
        LOGGER.info(
            "_collect_cylinder_form_data: シリンダー入力 data=%s",
            _summarize_for_log(cylinder_data),
        )
        return cylinder_data

    def fill_form(self, data: Dict[str, str]) -> None:
        for k, w in self.widgets.items():
            v = data.get(k, "")
            if isinstance(w, QPlainTextEdit):
                w.setPlainText(v)
            elif isinstance(w, QLineEdit):
                w.setText(v)

    def load_xlsm_path(self) -> Optional[str]:
        if os.path.exists(self.path_store):
            with open(self.path_store, "r", encoding="utf-8") as f:
                saved = f.read().strip()
            if saved:
                return saved
        selected = self.ask_xlsm_path()
        if selected:
            with open(self.path_store, "w", encoding="utf-8") as f:
                f.write(selected)
            return selected
        return None

    def ask_xlsm_path(self) -> Optional[str]:
        path, _ = QFileDialog.getOpenFileName(
            self, "xlsm を選んでください", "", "Excel マクロ有効ブック (*.xlsm)")
        return path or None

    @QtCore.Slot()
    def on_fetch(self):
        # 処理開始時刻を記録します
        self._fetch_start_time = time.perf_counter()

        # 入力欄から品目番号を取り出します
        item_no = ""
        if "品目番号" in self.widgets and isinstance(self.widgets["品目番号"], QLineEdit):
            item_no = self.widgets["品目番号"].text().strip()

        # 品目番号が空でないか確認します
        if not item_no:
            QMessageBox.warning(self, "入力エラー", "品目番号を入力してください。")
            elapsed = time.perf_counter() - self._fetch_start_time  # 処理にかかった時間を計算します
            print(f"[LOG] on_fetch: {elapsed:.3f} 秒")  # かかった時間を表示します
            return

        # 8桁の数字かどうか確認します
        if re.fullmatch(r"\d{8}", item_no) is None:
            QMessageBox.warning(self, "入力エラー", "品目番号は8桁の半角数字で入力してください。")
            elapsed = time.perf_counter() - self._fetch_start_time  # 処理にかかった時間を計算します
            print(f"[LOG] on_fetch: {elapsed:.3f} 秒")  # かかった時間を表示します
            return

        # 事前に作っておいたインデックスを取り出します
        sheet_index = self.record_index.get(self.excel_sheet)
        if not sheet_index:
            QMessageBox.warning(self, "データなし", "事前に読み込んだデータがありません。")
            elapsed = time.perf_counter() - self._fetch_start_time  # 処理にかかった時間を計算します
            print(f"[LOG] on_fetch: {elapsed:.3f} 秒")  # かかった時間を表示します
            return

        # 別スレッドで検索を行うワーカーを用意します
        self._fetch_worker = FetchWorker(sheet_index, item_no)
        # スレッド完了時に結果を受け取る関数をつなぎます
        self._fetch_worker.finished.connect(self._handle_fetch_result)
        # スレッドを開始します
        self._fetch_worker.start()

    def _handle_fetch_result(self, data: dict) -> None:
        """別スレッドでの検索結果を受け取り、画面を更新します"""
        # スレッドが終わったので参照を消しておきます
        self._fetch_worker = None

        # エラーがあればメッセージを表示します
        if "error" in data:
            QMessageBox.critical(self, "読み込みエラー", data["error"])
        else:
            rec = data.get("record")
            if rec is None:
                # 見つからなかった場合の処理です
                QMessageBox.information(self, "見つかりません", "新規入力できます。")
                self.on_clear(keep_item=True)
            else:
                # 見つかったレコードをフォームに反映します
                filtered = {k: rec.get(k, "") for k in self.widgets.keys()}
                self.fill_form(filtered)
                # シリンダー番号を画面に反映します
                zero_key = f"{to_full_width(0)}色目シリンダー"  # 「０色目シリンダー」の列名を作ります
                start = 0 if rec.get(zero_key, "") else 1  # 開始番号を決めます
                self.update_color_numbers(start)
                for idx, unit in enumerate(self.cylinder_units):
                    # 列名に使う数字を全角に変え、必要に応じてずらします
                    key = f"{to_full_width(idx + start)}色目シリンダー"
                    value = rec.get(key, "")
                    value = "" if value is None else str(value)
                    line = unit.cylinder_combo.lineEdit()
                    if line is not None:
                        line.setText(value)
                self.status.showMessage("事前データから読み込みました。", 3000)

        # 処理にかかった時間を計算し、ログに表示します
        elapsed = time.perf_counter() - self._fetch_start_time
        print(f"[LOG] on_fetch: {elapsed:.3f} 秒")  # かかった時間を表示します

    @QtCore.Slot()
    def on_save(self):
        data = self.collect_form_data()
        # 初学者向け説明：保存前に項目名を Excel と同じ形に整えます。
        normalized_data = normalize_form_keys(data)

        if not normalized_data.get("品目番号"):
            QMessageBox.warning(self, "入力エラー", "品目番号は必須です。")
            return
        if self.current_xlsm is None:
            QMessageBox.warning(
                self,
                "設定エラー",
                "データファイルが設定されていないため、保存できません。"
            )
            return

        # 初学者向け説明：ボタンに表示されている文字で、新規登録か上書きかを判断します。
        save_mode = "新規登録"
        if self.save_button is not None:
            save_mode = self.save_button.text().strip() or save_mode

        # 初学者向け説明：保存対象のパスやモード、入力値をログへ記録します。
        LOGGER.info(
            "on_save: 保存処理を開始します path=%s mode=%s data=%s",
            self.current_xlsm,
            save_mode,
            _summarize_for_log(normalized_data),
        )

        try:
            upsert_record_to_xlsm(self.current_xlsm, data, self.excel_sheet, save_mode)
            self.status.showMessage("Excel に保存しました。", 3000)
            QMessageBox.information(self, "保存", "保存が完了しました。")

            # 初学者向け説明：保存成功の事実をログへ残します。
            LOGGER.info("on_save: Excel への保存が完了しました")

            sheet_data = self.preloaded_data.get(self.excel_sheet)
            if sheet_data:
                # 初学者向け説明：見出しを空白のない名前にそろえて、項目名と合わせます。
                headers = [normalize_header_name(v) for v in sheet_data[0]]
                row = [normalized_data.get(h, "") for h in headers]
                item_key = normalized_data.get("品目番号", "")
                if "品目番号" in headers:
                    idx = headers.index("品目番号")
                    for i in range(1, len(sheet_data)):
                        cell = sheet_data[i][idx]
                        cell_text = str(cell) if cell is not None else ""
                        if cell_text == item_key:
                            sheet_data[i] = row
                            break
                    else:
                        sheet_data.append(row)
                # 辞書の内容も更新します
                sheet_index = self.record_index.setdefault(self.excel_sheet, {}).setdefault("品目番号", {})
                record_dict = {h: normalized_data.get(h, "") for h in headers if h}
                if item_key:
                    sheet_index[item_key] = record_dict

            self.update_button_states()
        except Exception as e:
            # 初学者向け説明：エラーの詳細をログへ記録し、利用者へも伝えます。
            LOGGER.exception("on_save: 保存処理でエラーが発生しました")
            QMessageBox.critical(self, "保存エラー", str(e))

    @QtCore.Slot()
    def on_clear(self, keep_item: bool = False):
        for k, w in self.widgets.items():
            if keep_item and k == "品目番号":
                continue
            if isinstance(w, (QLineEdit, QPlainTextEdit)):
                w.clear()

        if not keep_item:
            self._last_fetched_item = ""


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    layout_path = os.path.join(base, "layout.json")

    app = QApplication(sys.argv)
    apply_stylesheet(app, theme="light_blue.xml")
    # 無効状態の入力欄やボタンを灰色に、プルダウンの文字色を黒にします。
    app.setStyleSheet(app.styleSheet() + """
        QLineEdit:disabled,
        QPlainTextEdit:disabled,
        QPushButton:disabled {
            background-color: #E0E0E0;
            color: #9E9E9E;
        }
        QComboBox {
            color: #000000;
        }
    """)

    w = MainWindow(layout_path)
    w.showMaximized()  # ← 起動時に最大化
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
